import csv
import io
import logging
from datetime import datetime, timedelta
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.dependencies import get_current_admin, get_db
from app.repositories import reports_repository
from app.schemas.reports import (
    AnalisisResponse,
    OcupacionResponse,
    TaquillaResponse,
    VentaHorarioResponse,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/admin/reports", tags=["admin reports"])


PERIODOS_VALIDOS = ["hoy", "semana", "mes", "anio", "mes_anterior"]
TIPOS_EXPORT = ["taquilla", "ocupacion-salas", "ventas-horario", "analisis-peliculas", "detalle-compras"]

HEADERS_MAP = {
    "taquilla": ["ID", "Película", "Funciones", "Entradas", "Total Vendido", "Estado"],
    "ocupacion-salas": ["Sala", "Cine", "Capacidad", "Vendidos", "Ocupación %", "Formato"],
    "ventas-horario": ["Horario", "Transacciones", "Ingresos", "Tickets Vendidos"],
    "analisis-peliculas": ["Género", "Películas", "Funciones", "Ingresos", "% del Total"],
    "detalle-compras": ["ID", "Cliente", "Película", "Sala", "Boletos", "Confitería", "Total", "Fecha", "Estado"],
}

REPORT_FUNCTIONS = {
    "taquilla": reports_repository.get_taquilla_data,
    "ocupacion-salas": reports_repository.get_ocupacion_salas_data,
    "ventas-horario": reports_repository.get_ventas_horario_data,
    "analisis-peliculas": reports_repository.get_analisis_peliculas_data,
    "detalle-compras": reports_repository.get_detalle_compras_data,
}

FIELD_MAPS = {
    "taquilla": ["id", "titulo", "funciones", "entradas", "ingreso", "estado"],
    "ocupacion-salas": ["sala", "cine", "capacidad", "vendidos", "porcentaje", "formato"],
    "ventas-horario": ["horario", "ventas", "ingresos", "tickets"],
    "analisis-peliculas": ["genero", "peliculas", "funciones", "ingresos", "porcentaje"],
    "detalle-compras": ["id_transaccion", "cliente", "pelicula", "sala", "boletos", "confiteria", "total", "fecha", "estado"],
}

RESUMEN_MAP = {
    "taquilla": [
        ("Total Funciones", "total_funciones"),
        ("Total Entradas", "total_entradas"),
        ("Ingreso Bruto", "ingreso_bruto"),
        ("Ticket Promedio", "ticket_promedio"),
    ],
    "ocupacion-salas": [
        ("Total Salas", "total_salas"),
        ("Ocupación Promedio", "ocupacion_promedio"),
        ("Capacidad Total", "capacidad_total"),
    ],
    "ventas-horario": [
        ("Horario Pico", "horario_pico"),
        ("Ingreso Total", "ingreso_total"),
        ("Total Tickets", "total_tickets"),
    ],
    "analisis-peliculas": [
        ("Género Principal", "genero_principal"),
        ("Ingreso Total", "ingreso_total"),
        ("Total Películas", "total_peliculas"),
    ],
    "detalle-compras": [
        ("Total Transacciones", "total_transacciones"),
        ("Ingresos Totales", "total_ingresos"),
        ("Total Boletos", "total_boletos"),
        ("Total Snacks", "total_snacks"),
    ],
}


def _get_fecha_periodo_nombre(periodo: str):
    hoy = datetime.now()
    if periodo == "hoy":
        inicio = fin = hoy
    elif periodo == "semana":
        inicio = hoy - timedelta(days=7)
        fin = hoy
    elif periodo == "mes":
        inicio = hoy - timedelta(days=30)
        fin = hoy
    elif periodo == "anio":
        inicio = hoy - timedelta(days=365)
        fin = hoy
    elif periodo == "mes_anterior":
        inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        fin = inicio_mes
        inicio = (inicio_mes - timedelta(days=1)).replace(day=1)
    else:
        inicio = hoy - timedelta(days=30)
        fin = hoy
    return inicio, fin


@router.get("/taquilla", response_model=TaquillaResponse)
def report_taquilla(
    db: Annotated[Session, Depends(get_db)],
    _admin: Annotated[dict, Depends(get_current_admin)],
    periodo: Annotated[str, Query(pattern="^(hoy|semana|mes|anio|mes_anterior)$")] = "mes",
):
    return reports_repository.get_taquilla_data(db, periodo)


@router.get("/ocupacion-salas", response_model=OcupacionResponse)
def report_ocupacion_salas(
    db: Annotated[Session, Depends(get_db)],
    _admin: Annotated[dict, Depends(get_current_admin)],
    periodo: Annotated[str, Query(pattern="^(hoy|semana|mes|anio|mes_anterior)$")] = "mes",
):
    return reports_repository.get_ocupacion_salas_data(db, periodo)


@router.get("/ventas-horario", response_model=VentaHorarioResponse)
def report_ventas_horario(
    db: Annotated[Session, Depends(get_db)],
    _admin: Annotated[dict, Depends(get_current_admin)],
    periodo: Annotated[str, Query(pattern="^(hoy|semana|mes|anio|mes_anterior)$")] = "mes",
):
    return reports_repository.get_ventas_horario_data(db, periodo)


@router.get("/analisis-peliculas", response_model=AnalisisResponse)
def report_analisis_peliculas(
    db: Annotated[Session, Depends(get_db)],
    _admin: Annotated[dict, Depends(get_current_admin)],
    periodo: Annotated[str, Query(pattern="^(hoy|semana|mes|anio|mes_anterior)$")] = "mes",
):
    return reports_repository.get_analisis_peliculas_data(db, periodo)


@router.get("/detalle-compras")
def report_detalle_compras(
    db: Annotated[Session, Depends(get_db)],
    _admin: Annotated[dict, Depends(get_current_admin)],
    periodo: Annotated[str, Query(pattern="^(hoy|semana|mes|anio|mes_anterior)$")] = "mes",
):
    return reports_repository.get_detalle_compras_data(db, periodo)


@router.get("/export/excel")
def export_excel(
    db: Annotated[Session, Depends(get_db)],
    _admin: Annotated[dict, Depends(get_current_admin)],
    tipo: Annotated[str, Query(pattern="^(taquilla|ocupacion-salas|ventas-horario|analisis-peliculas|detalle-compras)$")],
    periodo: Annotated[str, Query(pattern="^(hoy|semana|mes|anio|mes_anterior)$")] = "mes",
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no está instalado")

    if tipo not in REPORT_FUNCTIONS:
        raise HTTPException(status_code=400, detail=f"Tipo inválido: {tipo}")

    result = REPORT_FUNCTIONS[tipo](db, periodo)
    data = result["data"]
    resumen = result["resumen"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    headers = HEADERS_MAP[tipo]
    fields = FIELD_MAPS[tipo]
    bold_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font

    for row_idx, item in enumerate(data, 2):
        for col_idx, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(field, ""))

    ws2 = wb.create_sheet("Resumen")
    ws2.cell(row=1, column=1, value="Métrica").font = bold_font
    ws2.cell(row=1, column=2, value="Valor").font = bold_font
    for row_idx, (label, key) in enumerate(RESUMEN_MAP[tipo], 2):
        ws2.cell(row=row_idx, column=1, value=label)
        ws2.cell(row=row_idx, column=2, value=resumen.get(key, ""))

    inicio, fin = _get_fecha_periodo_nombre(periodo)
    filename = f"reporte_{tipo}_{inicio.strftime('%Y-%m-%d')}_{fin.strftime('%Y-%m-%d')}.xlsx"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/csv")
def export_csv(
    db: Annotated[Session, Depends(get_db)],
    _admin: Annotated[dict, Depends(get_current_admin)],
    tipo: Annotated[str, Query(pattern="^(taquilla|ocupacion-salas|ventas-horario|analisis-peliculas|detalle-compras)$")],
    periodo: Annotated[str, Query(pattern="^(hoy|semana|mes|anio|mes_anterior)$")] = "mes",
):
    if tipo not in REPORT_FUNCTIONS:
        raise HTTPException(status_code=400, detail=f"Tipo inválido: {tipo}")

    result = REPORT_FUNCTIONS[tipo](db, periodo)
    data = result["data"]
    fields = FIELD_MAPS[tipo]

    buffer = io.StringIO()
    buffer.write("\ufeff")
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_NONNUMERIC)
    writer.writerow(HEADERS_MAP[tipo])

    for item in data:
        writer.writerow([str(item.get(f, "")) for f in fields])

    inicio, fin = _get_fecha_periodo_nombre(periodo)
    filename = f"reporte_{tipo}_{inicio.strftime('%Y-%m-%d')}_{fin.strftime('%Y-%m-%d')}.csv"

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/generados")
def reportes_generados(
    db: Annotated[Session, Depends(get_db)],
    _admin: Annotated[dict, Depends(get_current_admin)],
):
    row = reports_repository.get_reporte_contador(db)
    return {"count": row.count, "ultima_generacion": row.ultima_generacion.isoformat() if row.ultima_generacion else None}


@router.post("/generar")
def generar_reporte(
    db: Annotated[Session, Depends(get_db)],
    _admin: Annotated[dict, Depends(get_current_admin)],
):
    return reports_repository.incrementar_reporte_contador(db)
